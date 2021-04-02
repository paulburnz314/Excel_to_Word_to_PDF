import comtypes.client  # for saving a word file as a pdf
import time     # for delay to avoid errors
import openpyxl  # for getting data from excel
import pprint
from mailmerge import MailMerge     # for dumping values into word document where keys match merge fields
from input.list_dict_surcharges_21_input import monthly_surcharges
from hard_data import SURCHARGE_AMOUNT, THRESHOLD_AMOUNT, SIU_LIST, MERGE_FIELD_LABELS,\
    WRD_TO_PDF, PROJECT_FOLDER_OUTPUT


def replace_multiple(main_string, to_be_replaced, new_string):   # Iterate over the strings to be replaced
    for elem in to_be_replaced:   # Check if string is in the main string
        if elem in main_string:  # Replace the string
            main_string = main_string.replace(elem, new_string)
    return main_string


if __name__ == '__main__':
    row_number = 10
    number_of_months = 3  # plus one for cut off
    list_value = 0
    conc_over = 0
    load_over = 0
    surcharge_part = 0
    merge_match_results = []
    total_surcharge = 0
    users = SIU_LIST[0]
    # excel file that holds raw data exported from database
    WB_INPUT = openpyxl.load_workbook('./input/surcharge_data_from_database.xlsx')
    sheet1 = WB_INPUT['data']  # name of the sheet with the data
    sheet2 = WB_INPUT['contacts']  # name of the sheet with facility and contact information
    Contact1 = sheet2['A2':'A6']
    wb2 = openpyxl.Workbook()  # might as well dump the data processed into a new spreadsheet
    sheet3 = wb2.active
    sheet3.title = '2021_02_Summary'  # change as needed
    sheet3.append(MERGE_FIELD_LABELS)
    column_start = 3
    # which rows do you want to process, 12 months would be range(2,13)
    for row_num in range(row_number, row_number+number_of_months):
        month_year = sheet1.cell(row=row_num, column=1).value   # month and year always in column 1
        # get the month and year but remove the comma and
        month_year = replace_multiple(month_year, [','], '_')
        month_year = replace_multiple(month_year, [' '], '')
        for user in SIU_LIST:   # each user has six columns of data including flow, tss, cbod, nh3, tp and o&g
            doc_template = './input/Surcharge2021_Merge_Template.docx'
            document_1 = MailMerge(doc_template)
            flow = sheet1.cell(row=row_num, column=column_start - 1).value
            for column_num in range(1, 8):
                # fill the list merge_match_results with excel 'Contact2' sheet data
                merge_match_results.append(sheet2.cell(row=(SIU_LIST.index(user)) + 2, column=column_num).value)

            merge_match_results.append(month_year)
            flow = f'{flow:.6f}'    # format flow in millions of gallons and maintain 6 decimal places
            merge_match_results.append(str(flow))
            document_title = month_year + '_' + user + '_surcharges.docx'   # name of the word doc to be generated
            document_title2 = month_year + '_' + user + '_surcharges.pdf'   # name of the pdf to be generated
            # at start range is 3 to 8, note stops after column 7
            for column_num in range(column_start, column_start + 5):
                pollutant = sheet1.cell(row=row_num, column=column_num).value
                check_str = isinstance(pollutant, str)   # start checking for <, >, empty space or None values
                if check_str:
                    pollutant = replace_multiple(pollutant, ['>', '<'], '')
                if pollutant == ' ':
                    pollutant = 0
                elif pollutant is None:
                    pollutant = 0
                else:
                    pollutant = float(pollutant)  # data is cleaned up and converted to a number
                    if pollutant > 99.9:
                        pollutant = round(pollutant, 2)
                    else:
                        pollutant = round(pollutant, 3)
                # no surcharge if value is below threshold, no credit for low values either
                if pollutant > THRESHOLD_AMOUNT[list_value]:
                    flow = float(flow)
                    conc_over = round((pollutant - THRESHOLD_AMOUNT[list_value]), 3)
                    load_over = round(flow * 8.34 * conc_over, 3)
                    surcharge_part = round(SURCHARGE_AMOUNT[list_value] * load_over, 2)
                total_surcharge = total_surcharge + surcharge_part
                conc_over = f'{conc_over:.2f}'
                load_over = f'{load_over:.2f}'
                str_surcharge_part = f'{surcharge_part:.2f}'
                merge_match_results.append(str(pollutant))
                merge_match_results.append(conc_over)
                merge_match_results.append(load_over)
                merge_match_results.append(str_surcharge_part)
                list_value = list_value + 1
                conc_over = 0
                surcharge_part = 0
                load_over = 0
            str_total_surcharge = f'{total_surcharge:.2f}'
            merge_match_results.append(str_total_surcharge)
            sheet3.append(merge_match_results)      # dump data into summary excel sheet

            # combine the two lists into a dictionary so you can pass it to the word document
            merge_dict = {MERGE_FIELD_LABELS[i]: merge_match_results[i] for i in range(len(MERGE_FIELD_LABELS))}

            # append the combined keys and values to a list from a separate python file
            monthly_surcharges.append(dict(merge_dict))

            document_1.merge_templates([merge_dict], "page_break")
            # document_1.merge_pages([merge_dict])
            # NOTE if I was going to combine all the pages into
            # one document I would use above
            document_1.write("./output/" + document_title)
            document_1.close()
            # taking the just written document as in_file so it can be reopened and saved as a pdf
            in_file = PROJECT_FOLDER_OUTPUT + document_title
            out_file = PROJECT_FOLDER_OUTPUT + document_title2

            # creating COM object
            word = comtypes.client.CreateObject('Word.Application')
            # word.Visible = True  # kind of annoying having every document window pop up
            time.sleep(1)  # this slows down the file generation but avoids errors
            doc = word.Documents.Open(in_file)  # in_file is the word document just created
            doc.SaveAs(out_file, FileFormat=WRD_TO_PDF)  # out_file makes the pdf
            doc.Close()
            word.Quit()
            # word.Visible = False
            # set the column_start to grab the next set of columns on the row
            column_start = column_start + 6
            total_surcharge = 0     # reset values just in case previous data is still in there
            list_value = 0
            merge_match_results.clear()
            merge_dict.clear()
            # doc_template = ''

            wb2.save(filename='Surcharge_2021_02_Summary.xlsx')
            # after going through one user's pollutants, move on to the next,
            # once all are complete, drop to next row/month
        column_start = 3  # it is time to move down to the next row so set which column to start with
    format_list_dictionary = pprint.pformat(monthly_surcharges)
    with open("output/list_dict_surcharges_21_output.py", "w") as file:
        file.write(f"monthly_surcharges = {format_list_dictionary}")
